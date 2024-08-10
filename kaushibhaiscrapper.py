import signal
import sys
import time
import os
import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from datetime import datetime

# Define constants

# C:\Users\aaa\Downloads\chromedriver-win32\chromedriver-win32

SCROLL_COUNT = 1000  # Adjust the number of scrolls as needed
CHROME_DRIVER_PATH = r'C:\Users\aaa\Downloads\chromedriver-win32\chromedriver-win32\chromedriver.exe'  # Path to your ChromeDriver
PROGRESS_FILE = 'progress.xlsx'  # File to track progress
RESUME_FILE = 'resume_marker.txt'  # File to mark the last checkpoint

# Initialize the WebDriver with headless mode
def initialize_driver():
    options = Options()
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-popup-blocking')
    options.add_argument('--disable-infobars')
    options.add_argument('--headless')  # Headless mode for faster operation
    options.add_argument('--disable-gpu')

    driver = webdriver.Chrome(service=Service(CHROME_DRIVER_PATH), options=options)
    return driver

# Generate Google Maps URL based on query
def generate_google_maps_url(req_key):
    base_url = 'https://www.google.com/maps/search/'
    query = '+'.join(req_key.lower().split())  # Convert query to lowercase and replace spaces with '+'
    url = f'{base_url}{query}'
    return url

# Scroll the sidebar to the bottom
def scroll_sidebar_to_bottom(driver):
    sidebar_xpath = '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]'
    sidebar_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, sidebar_xpath)))
    driver.execute_script("arguments[0].scrollTo(0, arguments[0].scrollHeight);", sidebar_element)
    time.sleep(0.5)  # Reduced sleep time for faster operation

# Scrape names and numbers from the search results
def scrape_names_and_numbers(driver, field_name, location, existing_numbers):
    search_result_xpath = '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]'
    parent_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, search_result_xpath)))

    inner_div_class = 'Nv2PK tH5CWc THOPZb'
    nested_divs = parent_element.find_elements(By.XPATH, f".//div[contains(@class, '{inner_div_class}')]")
    results = []

    reached_css = 'HlvSq'
    reached_ele = parent_element.find_elements(By.XPATH, f".//span[contains(@class, '{reached_css}')]")
    is_reached = bool(reached_ele)

    for div in nested_divs:
        inner_css = 'NrDZNb'
        names = div.find_elements(By.XPATH, f".//div[contains(@class, '{inner_css}')]")
        num_css = 'UsdlK'
        nums = div.find_elements(By.XPATH, f".//span[contains(@class, '{num_css}')]")

        for name, num in zip(names, nums) if nums else zip(names, [''] * len(names)):
            name_content = name.text.strip()
            num_content = num.text.strip() if num else ''
            if num_content and num_content not in existing_numbers:
                results.append({'Name': name_content, 'Number': num_content, 'Field': field_name, 'Country': location})
                existing_numbers.add(num_content)

    return results, is_reached

# Write scraped data to Excel file
def write_to_excel(results, filename, city):
    folder = 'data'
    if not os.path.exists(folder):
        os.makedirs(folder)
    file_path = os.path.join(folder, filename)

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Data'
        ws.append(['Name', 'Number', 'Field', 'City'])

    ws = wb['Data'] if 'Data' in wb.sheetnames else wb.active

    # Find and delete existing data for the specified city
    max_row = ws.max_row
    rows_to_delete = [row for row in range(2, max_row + 1) if ws.cell(row=row, column=4).value == city]
    for row in reversed(rows_to_delete):
        ws.delete_rows(row, 1)

    # Write new data
    for entry in results:
        ws.append([entry['Name'], entry['Number'], entry['Field'], city])

    wb.save(file_path)
    return wb

# Retrieve cities and keywords from Excel file
def cities_keywords():
    file_path = 'keywords.xlsx'  # Replace with your file path
    xls = pd.ExcelFile(file_path)

    cities_df = pd.read_excel(xls, sheet_name='city')  # Sheet name for cities
    keywords_df = pd.read_excel(xls, sheet_name='keyword')  # Sheet name for keywords

    cities_list = cities_df.iloc[:, 0].tolist()
    keywords_list = keywords_df.iloc[:, 0].tolist()

    return cities_list, keywords_list

# Load progress from file
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        df = pd.read_excel(PROGRESS_FILE)
        return df.set_index(['Keyword', 'City']).to_dict(orient='index')
    return {}

# Save progress to file
def save_progress(progress):
    df = pd.DataFrame.from_dict(progress, orient='index').reset_index()
    df.to_excel(PROGRESS_FILE, index=False)


def load_last_checkpoint():
    if os.path.exists(RESUME_FILE):
        with open(RESUME_FILE, 'r') as file:
            line = file.readline().strip()
            if line:
                parts = line.split(',')
                if len(parts) == 2:
                    keyword, city = parts
                    return keyword, city
                else:
                    print(f"Warning: Unexpected format in {RESUME_FILE}. Expected 'keyword,city' but got: {line}")
    return None, None



# Save the last checkpoint
def save_last_checkpoint(keyword, city):
    with open(RESUME_FILE, 'w') as file:
        file.write(f'{keyword},{city}')

# Handle script interruption
def signal_handler(sig, frame):
    print('Script interrupted. Saving progress...')
    save_progress(progress)
    print('Progress saved. Exiting...')
    sys.exit(0)

# Register the signal handler
signal.signal(signal.SIGINT, signal_handler)

def main():
    cities, keywords = cities_keywords()
    driver = initialize_driver()
    total_rows = 0

    progress = load_progress()  # Load the progress
    last_keyword, last_city = load_last_checkpoint()  # Load the last checkpoint

    start_processing = False

    for keyword in keywords:
        if last_keyword and last_city:
            if keyword == last_keyword:
                start_processing = True
                cities = cities[cities.index(last_city):]
            else:
                continue
        
        for city in cities:
            if (keyword, city) in progress:
                print(f"Skipping already processed city '{city}' with keyword '{keyword}'.")
                continue
            
            location = city
            field_name = keyword

            url = generate_google_maps_url(f'{field_name} {location}')
            driver.get(url)

            filename = f'{keyword}.xlsx'
            existing_numbers = set()  # Set to track existing numbers and avoid duplicates

            try:
                for count in range(SCROLL_COUNT):
                    scroll_sidebar_to_bottom(driver)
                    results, is_reached = scrape_names_and_numbers(driver, field_name, location, existing_numbers)
                    current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                    print(f'{current_time} - scroll count: {count + 1} - Keyword: {keyword}, City: {city} > Results: {results}\n')

                    if results:
                        wb = write_to_excel(results, filename, city)
                        total_rows += len(results)
                        if is_reached:
                            break
                    else:
                        break

                # Mark the city and keyword as processed
                progress[(keyword, city)] = {'Keyword': keyword, 'City': city}
                save_progress(progress)
                save_last_checkpoint(keyword, city)

            except Exception as e:
                print(f"Error processing city '{city}' with keyword '{keyword}': {e}")

    driver.quit()
    print(f'Total data rows: {total_rows}')
    print('Data collection completed successfully.')

if __name__ == "__main__":
    main()
